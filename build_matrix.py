"""Build matrix XLSX from cleaned Dachser Document Intelligence JSON."""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from project_paths import OUTPUT_DIR, PROCESSING_DIR
from file_selection import prompt_input_json
from clean_di_json import write_cleaned_json
from postal_zones import (
    postal_zones_from_fields,
    postal_zones_txt_path_for_matrix,
    write_postal_zones_txt,
)


def _load_fields(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as handle:
        payload = json.load(handle)
    if not isinstance(payload, dict):
        raise ValueError("Expected JSON root object.")
    if "analyzeResult" in payload:
        from clean_di_json import clean_analyze_payload

        return clean_analyze_payload(payload)["documents"][0]["fields"]
    documents = payload.get("documents")
    if isinstance(documents, list) and documents:
        fields = documents[0].get("fields")
        if isinstance(fields, dict):
            return fields
    raise ValueError("JSON does not contain document fields.")

ORIGIN_COUNTRY = "IE"
CURRENCY = "EUR"

COST_NAME_ROW = 1
APPLY_IF_ROW = 2
RATE_BY_ROW = 3
BRACKET_ROW = 4
COLUMN_HEADER_ROW = 5
DATA_START_ROW = 6

HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
BOLD = Font(bold=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

SHIPMENT_HEADERS = [
    "Origin Country",
    "Destination Postal Code Zone",
    "Destination City",
]

PARCEL_PIECE_TYPES = ("KN", "KT", "DR")
PALLET_PIECE_TYPES = ("EW", "IP", "EU", "F", "IB", "DR")
PARCEL_BRACKET_LABELS: tuple[str, ...] = tuple(f"<={n}" for n in range(1, 11)) + (">10",)
PALLET_BRACKET_LABELS: tuple[str, ...] = tuple(f"<={n}" for n in range(1, 19))

DEFAULT_APPLY_IF = "Apply if: Applies in all items"


@dataclass(frozen=True)
class CostColumnSpec:
    bracket_label: str
    rate_unit: str = "Flat"


@dataclass
class CostBlock:
    title: str
    apply_if: str
    rate_by: str
    columns: list[CostColumnSpec] = field(default_factory=list)
    uses_shared_currency: bool = False


@dataclass
class MatrixRow:
    shipment: dict[str, Any]
    costs: dict[tuple[str, str], float | None] = field(default_factory=dict)


@dataclass
class SkippedItem:
    source: str
    label: str
    detail: str


def _parse_euro(value: object) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("€", "").replace(",", ".").strip()
    try:
        return float(text)
    except ValueError:
        return None


def _zone_letter(zone_header: str) -> str:
    first_line = zone_header.split("\n", 1)[0].strip()
    match = re.search(r"Zone\s+([A-E])", first_line, re.IGNORECASE)
    if match:
        return match.group(1).upper()
    return first_line


def _destination_postal_zone(zone_header: str) -> str:
    letter = _zone_letter(zone_header)
    return f"IE Zone {letter}"


def _main_cost_rates(main_costs: list[dict[str, Any]]) -> dict[int, dict[str, float]]:
    """Map pallet count 1-18 -> zone column -> rate."""
    rates: dict[int, dict[str, float]] = {}
    for row in main_costs:
        pallet_raw = str(row.get("Pallets", "")).strip()
        if not pallet_raw.isdigit():
            continue
        pallet_qty = int(pallet_raw)
        if pallet_qty < 1 or pallet_qty > 18:
            continue
        zone_rates: dict[str, float] = {}
        for zone_key, raw in row.items():
            if not zone_key.startswith("Zone"):
                continue
            amount = _parse_euro(raw)
            if amount is not None:
                zone_rates[zone_key] = amount
        if zone_rates:
            rates[pallet_qty] = zone_rates
    return rates


def _parcel_source_pallet(bracket_label: str) -> int:
    if bracket_label == ">10":
        return 1
    if bracket_label.startswith("<="):
        return int(bracket_label[2:])
    raise ValueError(f"Unknown parcel bracket label: {bracket_label}")


def _pallet_source_pallet(bracket_label: str) -> int:
    if bracket_label.startswith("<="):
        return int(bracket_label[2:])
    raise ValueError(f"Unknown pallet bracket label: {bracket_label}")


def _find_additional_row(
    rows: list[dict[str, Any]], name_fragment: str
) -> dict[str, Any] | None:
    fragment = name_fragment.casefold()
    for row in rows:
        name = str(row.get("Name", "")).casefold()
        if fragment in name:
            return row
    return None


def _zone_mapped_value(
    zone_letter: str,
    *,
    dublin_value: float | None,
    country_value: float | None,
) -> float | None:
    if zone_letter == "A":
        return dublin_value
    return country_value


@dataclass(frozen=True)
class ProductSpec:
    code: str
    parcel: bool


def _kind_label(*, parcel: bool) -> str:
    return "Parcel" if parcel else "Pallet"


def _iter_product_specs() -> list[ProductSpec]:
    specs: list[ProductSpec] = [
        ProductSpec(code=code, parcel=True) for code in PARCEL_PIECE_TYPES
    ]
    specs.extend(ProductSpec(code=code, parcel=False) for code in PALLET_PIECE_TYPES)
    return specs


def _bracket_labels_for_product(spec: ProductSpec) -> tuple[str, ...]:
    return PARCEL_BRACKET_LABELS if spec.parcel else PALLET_BRACKET_LABELS


def _source_pallet_qty(spec: ProductSpec, bracket_label: str) -> int:
    if spec.parcel:
        return _parcel_source_pallet(bracket_label)
    return _pallet_source_pallet(bracket_label)


def _transport_title(spec: ProductSpec) -> str:
    kind = _kind_label(parcel=spec.parcel)
    return f"Transport cost ({kind} {spec.code})"


def _second_delivery_title(spec: ProductSpec) -> str:
    kind = _kind_label(parcel=spec.parcel)
    return f"Second Delivery ({kind} {spec.code})"


def _product_rate_by(spec: ProductSpec) -> str:
    if spec.parcel:
        return "Rate by: Per pallet (1-10; >10 uses bracket 1)"
    return "Rate by: Per pallet (1-18)"


def _product_bracket_specs(spec: ProductSpec) -> list[CostColumnSpec]:
    return [
        CostColumnSpec(bracket_label=label, rate_unit="Flat")
        for label in _bracket_labels_for_product(spec)
    ]


def _product_cost_block(spec: ProductSpec, *, second_delivery: bool) -> CostBlock:
    title = _second_delivery_title(spec) if second_delivery else _transport_title(spec)
    return CostBlock(
        title=title,
        apply_if=f"Apply if: Piece type equals {spec.code}",
        rate_by=_product_rate_by(spec),
        columns=_product_bracket_specs(spec),
        uses_shared_currency=True,
    )


def _build_product_cost_blocks() -> list[CostBlock]:
    blocks: list[CostBlock] = []
    for spec in _iter_product_specs():
        blocks.append(_product_cost_block(spec, second_delivery=False))
        blocks.append(_product_cost_block(spec, second_delivery=True))
    return blocks


def _ancillary_cost_blocks() -> list[CostBlock]:
    return [
        CostBlock(
            title="Transport cost (Empty IBC)",
            apply_if=DEFAULT_APPLY_IF,
            rate_by="Rate by: Per IBC",
            columns=[CostColumnSpec(bracket_label="Flat", rate_unit="Flat")],
        ),
        CostBlock(
            title="Pump Surcharge",
            apply_if=DEFAULT_APPLY_IF,
            rate_by="Rate by: Per tank/IBC",
            columns=[
                CostColumnSpec(bracket_label="<=1", rate_unit="Flat"),
                CostColumnSpec(bracket_label=">1", rate_unit="Flat"),
            ],
            uses_shared_currency=True,
        ),
        CostBlock(
            title="AM / PM Deliveries",
            apply_if=DEFAULT_APPLY_IF,
            rate_by="Rate by: Per delivery",
            columns=[CostColumnSpec(bracket_label="Flat", rate_unit="Flat")],
        ),
        CostBlock(
            title="Timed delivery",
            apply_if=DEFAULT_APPLY_IF,
            rate_by="Rate by: Per delivery",
            columns=[CostColumnSpec(bracket_label="Flat", rate_unit="Flat")],
        ),
        CostBlock(
            title="Booking In",
            apply_if=DEFAULT_APPLY_IF,
            rate_by="Rate by: Per delivery",
            columns=[CostColumnSpec(bracket_label="Flat", rate_unit="Flat")],
        ),
    ]


def cost_key(block: CostBlock, spec: CostColumnSpec) -> tuple[str, str]:
    return (block.title, spec.bracket_label)


def block_column_width(block: CostBlock) -> int:
    if block.uses_shared_currency:
        return 1 + len(block.columns)
    return 2 * len(block.columns)


def _has_cost(value: float | None) -> bool:
    return value is not None


def build_matrix(
    fields: dict[str, Any],
) -> tuple[list[MatrixRow], list[CostBlock], list[SkippedItem]]:
    main_costs = fields.get("MainCosts") or []
    additional = fields.get("AdditionalCosts") or []
    additional2 = fields.get("AdditionalCosts2") or []
    skipped: list[SkippedItem] = []

    if not main_costs:
        raise ValueError("MainCosts is missing or empty.")

    header_row = main_costs[0]
    zone_keys = [key for key in header_row if key.startswith("Zone")]
    pallet_rates = _main_cost_rates(main_costs)

    second_delivery_row = _find_additional_row(
        additional2, "2nd Delivery Charge"
    )
    timed_row = _find_additional_row(additional2, "Time specific")
    booking_row = _find_additional_row(additional2, "Book in charge")
    empty_ibc_row = _find_additional_row(additional, "Empty IBC")
    pump_first_row = _find_additional_row(additional, "First Tank")
    pump_extra_row = _find_additional_row(additional, "Therafter")

    dublin_timed = _parse_euro(timed_row.get("Zone1") if timed_row else None)
    country_timed = _parse_euro(timed_row.get("Zone2") if timed_row else None)
    am_pm_dublin = 20.0
    am_pm_country = 30.0
    booking_flat = _parse_euro(booking_row.get("Charge") if booking_row else None)
    empty_ibc = _parse_euro(empty_ibc_row.get("Price") if empty_ibc_row else None)
    pump_first = _parse_euro(pump_first_row.get("Price") if pump_first_row else None)
    pump_extra = _parse_euro(pump_extra_row.get("Price") if pump_extra_row else None)

    product_blocks = _build_product_cost_blocks()
    ancillary_blocks = _ancillary_cost_blocks()
    cost_blocks = product_blocks + ancillary_blocks

    matrix_rows: list[MatrixRow] = []
    for zone_key in zone_keys:
        zone_header = str(header_row.get(zone_key, ""))
        zone_letter = _zone_letter(zone_header)

        shipment = {
            "Origin Country": ORIGIN_COUNTRY,
            "Destination Postal Code Zone": _destination_postal_zone(zone_header),
            "Destination City": None,
        }
        costs: dict[tuple[str, str], float | None] = {}

        for spec in _iter_product_specs():
            transport_title = _transport_title(spec)
            second_title = _second_delivery_title(spec)
            for bracket_label in _bracket_labels_for_product(spec):
                source_pallet = _source_pallet_qty(spec, bracket_label)
                rate = pallet_rates.get(source_pallet, {}).get(zone_key)
                costs[(transport_title, bracket_label)] = rate
                costs[(second_title, bracket_label)] = rate

        empty_block = ancillary_blocks[0]
        costs[cost_key(empty_block, empty_block.columns[0])] = empty_ibc

        pump_block = ancillary_blocks[1]
        costs[cost_key(pump_block, pump_block.columns[0])] = pump_first
        costs[cost_key(pump_block, pump_block.columns[1])] = pump_extra

        am_pm_block = ancillary_blocks[2]
        costs[cost_key(am_pm_block, am_pm_block.columns[0])] = _zone_mapped_value(
            zone_letter,
            dublin_value=am_pm_dublin,
            country_value=am_pm_country,
        )

        timed_block = ancillary_blocks[3]
        costs[cost_key(timed_block, timed_block.columns[0])] = _zone_mapped_value(
            zone_letter,
            dublin_value=dublin_timed,
            country_value=country_timed,
        )

        booking_block = ancillary_blocks[4]
        costs[cost_key(booking_block, booking_block.columns[0])] = booking_flat

        matrix_rows.append(MatrixRow(shipment=shipment, costs=costs))

    for row in main_costs:
        pallets = str(row.get("Pallets", "")).strip()
        if pallets in {"Pallet Qty"}:
            continue
        if pallets.isdigit() and 1 <= int(pallets) <= 18:
            continue
        skipped.append(
            SkippedItem(
                source="MainCosts",
                label=pallets or "(blank)",
                detail=json.dumps(row, ensure_ascii=False),
            )
        )

    for row in additional:
        name = str(row.get("Name", "")).strip()
        if empty_ibc_row is not None and row is empty_ibc_row:
            continue
        if pump_first_row is not None and row is pump_first_row:
            continue
        if pump_extra_row is not None and row is pump_extra_row:
            continue
        skipped.append(
            SkippedItem(
                source="AdditionalCosts",
                label=name,
                detail=json.dumps(row, ensure_ascii=False),
            )
        )

    for row in additional2:
        name = str(row.get("Name", "")).strip()
        if row is second_delivery_row:
            skipped.append(
                SkippedItem(
                    source="AdditionalCosts2",
                    label=name,
                    detail=(
                        "Not used: Second Delivery matrix columns mirror "
                        "Transport cost from MainCosts."
                    ),
                )
            )
            continue
        if row is timed_row or row is booking_row:
            continue
        if "Special Requests" in name:
            skipped.append(
                SkippedItem(
                    source="AdditionalCosts2",
                    label=name,
                    detail="Header row only (Dublin -> Zone A mapping).",
                )
            )
            continue
        if name.startswith("AM / PM"):
            skipped.append(
                SkippedItem(
                    source="AdditionalCosts2",
                    label=name,
                    detail="OCR row discarded; matrix uses AM / PM Deliveries €20 (Zone A) / €30 (other zones).",
                )
            )
            continue
        skipped.append(
            SkippedItem(
                source="AdditionalCosts2",
                label=name,
                detail=json.dumps(row, ensure_ascii=False),
            )
        )

    return matrix_rows, cost_blocks, skipped


def write_rates_sheet(
    worksheet,
    matrix_rows: list[MatrixRow],
    cost_blocks: list[CostBlock],
) -> None:
    shipment_count = len(SHIPMENT_HEADERS)

    def write_merged_row(row_index: int, values: list[str]) -> None:
        column_index = shipment_count + 1
        for block_index, block in enumerate(cost_blocks):
            width = block_column_width(block)
            value = values[block_index] if block_index < len(values) else ""
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            cell.font = BOLD
            cell.fill = HEADER_FILL
            cell.alignment = LEFT
            if width > 1:
                worksheet.merge_cells(
                    start_row=row_index,
                    start_column=column_index,
                    end_row=row_index,
                    end_column=column_index + width - 1,
                )
            column_index += width

    write_merged_row(COST_NAME_ROW, [block.title for block in cost_blocks])
    write_merged_row(APPLY_IF_ROW, [block.apply_if for block in cost_blocks])
    write_merged_row(RATE_BY_ROW, [block.rate_by for block in cost_blocks])

    for col_idx, header in enumerate(SHIPMENT_HEADERS, start=1):
        cell = worksheet.cell(row=COLUMN_HEADER_ROW, column=col_idx, value=header)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = LEFT

    column_index = shipment_count + 1
    for block in cost_blocks:
        if block.uses_shared_currency:
            currency_cell = worksheet.cell(row=BRACKET_ROW, column=column_index)
            currency_cell.fill = HEADER_FILL
            worksheet.cell(row=COLUMN_HEADER_ROW, column=column_index, value=CURRENCY)
            for offset, spec in enumerate(block.columns):
                spec_col = column_index + 1 + offset
                bracket_cell = worksheet.cell(
                    row=BRACKET_ROW, column=spec_col, value=spec.bracket_label
                )
                bracket_cell.font = BOLD
                bracket_cell.fill = HEADER_FILL
                bracket_cell.alignment = CENTER
                unit_cell = worksheet.cell(
                    row=COLUMN_HEADER_ROW, column=spec_col, value=spec.rate_unit
                )
                unit_cell.font = BOLD
                unit_cell.fill = HEADER_FILL
                unit_cell.alignment = CENTER
            column_index += block_column_width(block)
            continue

        for spec in block.columns:
            bracket_cell = worksheet.cell(
                row=BRACKET_ROW, column=column_index, value=spec.bracket_label
            )
            bracket_cell.font = BOLD
            bracket_cell.fill = HEADER_FILL
            bracket_cell.alignment = CENTER
            worksheet.merge_cells(
                start_row=BRACKET_ROW,
                start_column=column_index,
                end_row=BRACKET_ROW,
                end_column=column_index + 1,
            )
            worksheet.cell(row=COLUMN_HEADER_ROW, column=column_index, value=CURRENCY)
            unit_cell = worksheet.cell(
                row=COLUMN_HEADER_ROW, column=column_index + 1, value=spec.rate_unit
            )
            unit_cell.font = BOLD
            unit_cell.fill = HEADER_FILL
            unit_cell.alignment = CENTER
            column_index += 2

    for row_offset, matrix_row in enumerate(matrix_rows):
        excel_row = DATA_START_ROW + row_offset
        for col_idx, header in enumerate(SHIPMENT_HEADERS, start=1):
            worksheet.cell(
                row=excel_row, column=col_idx, value=matrix_row.shipment.get(header)
            )

        column_index = shipment_count + 1
        for block in cost_blocks:
            if block.uses_shared_currency:
                worksheet.cell(row=excel_row, column=column_index, value=CURRENCY)
                for offset, spec in enumerate(block.columns):
                    spec_col = column_index + 1 + offset
                    value = matrix_row.costs.get(cost_key(block, spec))
                    if _has_cost(value):
                        cell = worksheet.cell(row=excel_row, column=spec_col, value=value)
                        cell.number_format = "0.00"
                column_index += block_column_width(block)
                continue

            for spec in block.columns:
                value = matrix_row.costs.get(cost_key(block, spec))
                if _has_cost(value):
                    worksheet.cell(row=excel_row, column=column_index, value=CURRENCY)
                    cell = worksheet.cell(row=excel_row, column=column_index + 1, value=value)
                    cell.number_format = "0.00"
                column_index += 2

    for col_idx in range(1, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = 18


def write_skipped_sheet(worksheet, items: list[SkippedItem]) -> None:
    headers = ["Source", "Label", "Detail"]
    for col_idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        cell.font = BOLD
        cell.fill = HEADER_FILL
    for row_idx, item in enumerate(items, start=2):
        worksheet.cell(row=row_idx, column=1, value=item.source)
        worksheet.cell(row=row_idx, column=2, value=item.label)
        worksheet.cell(row=row_idx, column=3, value=item.detail)
    for col_idx in range(1, 4):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = 36


def build_matrix_workbook(cleaned_path: Path, output_path: Path | None = None) -> Path:
    fields = _load_fields(cleaned_path)

    matrix_rows, cost_blocks, skipped = build_matrix(fields)
    if not matrix_rows:
        raise ValueError("No matrix rows produced.")

    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        stem = cleaned_path.stem.replace(".cleaned", "")
        output_path = OUTPUT_DIR / f"matrix_{stem}_{timestamp}.xlsx"

    workbook = Workbook()
    rates_sheet = workbook.active
    rates_sheet.title = "Rates"
    write_rates_sheet(rates_sheet, matrix_rows, cost_blocks)

    skipped_sheet = workbook.create_sheet("did not added")
    write_skipped_sheet(skipped_sheet, skipped)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    zones = postal_zones_from_fields(fields)
    zones_txt_path = postal_zones_txt_path_for_matrix(output_path)
    write_postal_zones_txt(zones, zones_txt_path)

    return output_path


def main() -> None:
    source = prompt_input_json()
    cleaned_path = PROCESSING_DIR / f"{source.stem}.cleaned.json"
    if not cleaned_path.is_file():
        print(f"Cleaning {source.name}...")
        write_cleaned_json(source, cleaned_path)
    else:
        print(f"Using cleaned file {cleaned_path.name}")
    output_path = build_matrix_workbook(cleaned_path)
    zones_txt = postal_zones_txt_path_for_matrix(output_path)
    print(f"Matrix written to {output_path}")
    print(f"Postal zones written to {zones_txt}")


if __name__ == "__main__":
    main()
